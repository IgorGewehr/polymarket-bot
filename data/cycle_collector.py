"""
Coletor de dados por ciclo — registra snapshots a cada 30s e exporta para Excel.

Snapshots: 4:50, 4:30, 4:00, 3:30, 3:00, 2:30, 2:00, 1:30, 1:00, 0:30
Para cada snapshot: delta, direção, preço YES, retorno hipotético de $1.
Após resolução: calcula retorno real de $1 em cada time slot.
"""
import time
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
import structlog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

log = structlog.get_logger()

# Time slots em segundos restantes — 4:50 = 290s, 4:30 = 270s, etc.
SNAPSHOT_SLOTS = [
    ("4:50", 290),
    ("4:30", 270),
    ("4:00", 240),
    ("3:30", 210),
    ("3:00", 180),
    ("2:30", 150),
    ("2:00", 120),
    ("1:30", 90),
    ("1:00", 60),
    ("0:30", 30),
]

EXCEL_PATH = Path("./data/cycle_data.xlsx")


@dataclass
class Snapshot:
    """Dados capturados num instante do ciclo."""
    time_slot: str
    time_remaining: float
    delta: float
    direction: str        # "Up" ou "Down" (inferido do delta)
    yes_price: float
    no_price: float
    btc_price: float


@dataclass
class CycleRecord:
    """Dados completos de um ciclo de 5 minutos."""
    cycle_start: float = 0.0
    market_id: str = ""
    market_question: str = ""
    snapshots: dict[str, Snapshot] = field(default_factory=dict)
    result: str = ""            # "YES" ou "NO" (quem ganhou)
    final_yes_price: float = 0.0
    traded: bool = False        # Se o bot abriu posição neste ciclo
    trade_direction: str = ""   # Direção do trade real
    trade_size: float = 0.0     # Sizing real
    trade_entry_price: float = 0.0
    trade_pnl: float = 0.0

    def calc_return_1usd(self, slot: str) -> float | None:
        """Calcula retorno de $1 apostado no slot dado, após resolução."""
        snap = self.snapshots.get(slot)
        if not snap or not self.result:
            return None

        direction = snap.direction
        if direction == "Up":
            price = snap.yes_price
            if price <= 0 or price >= 1:
                return None
            if self.result == "YES":
                return round((1.0 / price) - 1.0, 4)  # Lucro
            else:
                return -1.0  # Perdeu $1
        else:  # Down
            price = snap.no_price
            if price <= 0 or price >= 1:
                return None
            if self.result == "NO":
                return round((1.0 / price) - 1.0, 4)
            else:
                return -1.0


class CycleCollector:
    """Coleta dados de cada ciclo e exporta para Excel."""

    def __init__(self, excel_path: Path = EXCEL_PATH):
        self.excel_path = excel_path
        self.current: CycleRecord | None = None
        self._captured_slots: set[str] = set()

    def start_cycle(self, market_id: str, question: str):
        """Inicia coleta para um novo ciclo."""
        self.current = CycleRecord(
            cycle_start=time.time(),
            market_id=market_id,
            market_question=question,
        )
        self._captured_slots.clear()
        log.debug("cycle_collector_start", market=question[:50])

    def capture_snapshot(
        self,
        time_remaining: float,
        delta: float,
        yes_price: float,
        btc_price: float,
    ):
        """Captura snapshot se estamos num time slot relevante."""
        if not self.current or yes_price <= 0:
            return

        for slot_name, slot_seconds in SNAPSHOT_SLOTS:
            if slot_name in self._captured_slots:
                continue
            # Captura quando time_remaining cruza o threshold (margem de 5s)
            if slot_seconds - 5 <= time_remaining <= slot_seconds + 5:
                direction = "Up" if delta >= 0 else "Down"
                snap = Snapshot(
                    time_slot=slot_name,
                    time_remaining=time_remaining,
                    delta=round(delta, 2),
                    direction=direction,
                    yes_price=round(yes_price, 4),
                    no_price=round(1.0 - yes_price, 4),
                    btc_price=round(btc_price, 2),
                )
                self.current.snapshots[slot_name] = snap
                self._captured_slots.add(slot_name)
                log.debug("snapshot_captured",
                          slot=slot_name,
                          delta=snap.delta,
                          dir=direction,
                          yes=snap.yes_price)
                break  # Um por iteração

    def record_trade(self, direction: str, size: float, entry_price: float):
        """Registra que o bot fez um trade neste ciclo."""
        if self.current:
            self.current.traded = True
            self.current.trade_direction = direction
            self.current.trade_size = size
            self.current.trade_entry_price = entry_price

    def end_cycle(self, final_yes_price: float, pnl: float = 0.0):
        """Finaliza o ciclo, calcula resultado e salva no Excel."""
        if not self.current:
            return

        # Determinar resultado
        if final_yes_price > 0.5:
            self.current.result = "YES"
        else:
            self.current.result = "NO"

        self.current.final_yes_price = final_yes_price
        self.current.trade_pnl = pnl

        self._save_to_excel(self.current)
        log.info("cycle_saved_to_excel",
                 market=self.current.market_question[:40],
                 result=self.current.result,
                 snapshots=len(self.current.snapshots),
                 traded=self.current.traded)

        self.current = None
        self._captured_slots.clear()

    def _save_to_excel(self, record: CycleRecord):
        """Salva o registro do ciclo no arquivo Excel."""
        if self.excel_path.exists():
            wb = load_workbook(self.excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ciclos BTC 5min"
            self._write_header(ws)

        row = ws.max_row + 1
        self._write_row(ws, row, record)
        wb.save(self.excel_path)

    def _write_header(self, ws):
        """Escreve o cabeçalho da planilha."""
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(bold=True, size=10, color="FFFFFF")

        headers = [
            "Data/Hora",
            "Market ID",
            "Mercado",
            "Resultado",
            "Apostou?",
            "Dir. Trade",
            "Sizing",
            "Preço Entrada",
            "PnL Trade",
        ]

        # Para cada time slot: Delta, Direção, Preço YES, Retorno $1
        for slot_name, _ in SNAPSHOT_SLOTS:
            headers.extend([
                f"Delta {slot_name}",
                f"Dir {slot_name}",
                f"YES {slot_name}",
                f"$1 Ret {slot_name}",
            ])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Ajustar largura das colunas
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 10
        for col_idx in range(5, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12

    def _write_row(self, ws, row: int, record: CycleRecord):
        """Escreve uma linha com todos os dados do ciclo."""
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        dt = datetime.fromtimestamp(record.cycle_start).strftime("%Y-%m-%d %H:%M:%S")

        base = [
            dt,
            record.market_id[:12] + "..." if len(record.market_id) > 12 else record.market_id,
            record.market_question[:60],
            record.result,
            "SIM" if record.traded else "NAO",
            record.trade_direction if record.traded else "-",
            f"${record.trade_size}" if record.traded else "-",
            f"{record.trade_entry_price:.4f}" if record.traded else "-",
            f"${record.trade_pnl:+.2f}" if record.traded else "-",
        ]

        # Dados de cada time slot
        for slot_name, _ in SNAPSHOT_SLOTS:
            snap = record.snapshots.get(slot_name)
            if snap:
                ret = record.calc_return_1usd(slot_name)
                ret_str = f"${ret:+.4f}" if ret is not None else "-"
                base.extend([
                    snap.delta,
                    snap.direction,
                    snap.yes_price,
                    ret_str,
                ])
            else:
                base.extend(["-", "-", "-", "-"])

        for col, value in enumerate(base, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Colorir resultado
        result_cell = ws.cell(row=row, column=4)
        if record.result == "YES":
            result_cell.fill = green_fill
        else:
            result_cell.fill = red_fill

        # Colorir PnL do trade
        if record.traded:
            pnl_cell = ws.cell(row=row, column=9)
            if record.trade_pnl > 0:
                pnl_cell.fill = green_fill
            elif record.trade_pnl < 0:
                pnl_cell.fill = red_fill

        # Colorir retornos hipotéticos de $1
        for i, (slot_name, _) in enumerate(SNAPSHOT_SLOTS):
            ret = record.calc_return_1usd(slot_name)
            if ret is not None:
                ret_col = 9 + (i * 4) + 4  # Coluna do retorno
                ret_cell = ws.cell(row=row, column=ret_col)
                if ret > 0:
                    ret_cell.fill = green_fill
                elif ret < 0:
                    ret_cell.fill = red_fill
