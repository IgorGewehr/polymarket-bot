"""
API REST para o dashboard do bot.
Roda como servidor FastAPI em paralelo ao engine.
"""
import time
from datetime import datetime
from fastapi import FastAPI
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pathlib import Path

app = FastAPI(title="Polymarket Bot Dashboard", docs_url=None, redoc_url=None)

# Referência ao engine (injetada no startup)
_engine = None


def set_engine(engine):
    global _engine
    _engine = engine


# ── Static files ───────────────────────────────────────────
DASHBOARD_DIR = Path(__file__).parent

app.mount("/static", StaticFiles(directory=DASHBOARD_DIR), name="static")


@app.get("/")
async def index():
    return FileResponse(DASHBOARD_DIR / "index.html")


# ── API Endpoints ──────────────────────────────────────────

@app.get("/api/status")
async def get_status():
    if not _engine:
        return {"error": "Engine not initialized"}

    rm = _engine.risk_manager
    pos = _engine.position
    market = _engine.current_market

    # PnL desde o reset (13:07 BRT = 16:07 UTC, banca $39)
    PNL_RESET_TS = 1776436060  # 2026-04-17T16:07:40Z
    daily_pnl = 0.0
    daily_trades = 0
    daily_wins = 0
    daily_losses = 0
    daily_locks = 0
    try:
        today_ts = PNL_RESET_TS
        row = _engine.storage.conn.execute("""
            SELECT COALESCE(SUM(pnl), 0),
                   COUNT(*),
                   COALESCE(SUM(CASE WHEN pnl > 0 THEN 1 ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN pnl < 0 THEN 1 ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN result LIKE '%LOCK%' THEN 1 ELSE 0 END), 0)
            FROM trades
            WHERE timestamp >= ? AND result IS NOT NULL
        """, [today_ts]).fetchone()
        if row:
            daily_pnl, daily_trades, daily_wins, daily_losses, daily_locks = row
    except Exception:
        pass

    return {
        "running": _engine.running,
        "dry_run": False,
        "timestamp": time.time(),
        "market": {
            "question": market.get("question", "?")[:80] if market else None,
            "time_remaining": round(_engine._get_time_remaining(market), 1) if market else 0,
        } if market else None,
        "position": {
            "yes_entry": round(pos.yes_entry, 4) if pos.yes_entry else None,
            "no_entry": round(pos.no_entry, 4) if pos.no_entry else None,
            "yes_shares": pos.yes_shares,
            "no_shares": pos.no_shares,
            "locked": pos.is_locked,
            "locked_profit": round(pos.locked_profit, 2) if pos.is_locked else None,
            "total_cost": round(pos.total_cost, 2),
        } if pos else None,
        "risk": rm.get_summary(),
        "daily": {
            "pnl": round(daily_pnl, 2),
            "trades": daily_trades,
            "wins": daily_wins,
            "losses": daily_losses,
            "locks": daily_locks,
        },
        "prices": {
            "yes": round(_engine.poly_feed.yes_price, 4),
            "no": round(_engine.poly_feed.no_price, 4),
            "btc": round(_engine.btc_feed.last_price, 2),
        },
    }


@app.get("/api/trades")
async def get_trades():
    if not _engine:
        return []
    try:
        trades = _engine.storage.get_recent_trades(100)
        return trades
    except Exception:
        return []


@app.get("/api/pnl")
async def get_pnl():
    """PnL cumulativo por trade para gráfico."""
    if not _engine:
        return []
    try:
        result = _engine.storage.conn.execute("""
            SELECT timestamp, pnl, direction, bet_size, result,
                   SUM(pnl) OVER (ORDER BY timestamp) as cumulative_pnl
            FROM trades
            WHERE result IS NOT NULL
            ORDER BY timestamp
        """).fetchall()
        cols = ["timestamp", "pnl", "direction", "bet_size", "result", "cumulative_pnl"]
        return [dict(zip(cols, row)) for row in result]
    except Exception:
        return []


@app.get("/api/cycles")
async def get_cycles():
    """Dados dos ciclos do Excel para tabela."""
    try:
        from openpyxl import load_workbook
        path = Path("./data/cycle_data.xlsx")
        if not path.exists():
            return []
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        wb.close()
        return rows[-50:]  # Últimos 50 ciclos
    except Exception:
        return []


@app.get("/api/daily")
async def get_daily():
    """Stats diárias."""
    if not _engine:
        return []
    try:
        result = _engine.storage.conn.execute("""
            SELECT * FROM daily_stats ORDER BY date DESC LIMIT 30
        """).fetchall()
        if not result:
            return []
        cols = [d[0] for d in _engine.storage.conn.description]
        return [dict(zip(cols, row)) for row in result]
    except Exception:
        return []


@app.post("/api/pause")
async def pause_bot():
    if _engine:
        _engine.running = False
        return {"status": "paused"}
    return {"error": "Engine not initialized"}


@app.post("/api/resume")
async def resume_bot():
    if _engine:
        _engine.running = True
        return {"status": "resumed"}
    return {"error": "Engine not initialized"}


@app.post("/api/unlock")
async def unlock_bot():
    """Destrava o bot após lock por 3 losses consecutivos."""
    if _engine:
        _engine.risk_manager.unlock()
        return {"status": "unlocked", "msg": "Bot destravado, pode operar novamente"}
    return {"error": "Engine not initialized"}
